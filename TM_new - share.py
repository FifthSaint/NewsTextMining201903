# -*- coding: utf-8 -*-
"""
Created on Fri Jul 19 15:55:24 2019

@author: sage5th
py37
for 세바데 토픽 모델링 '동성애'

based on "신년기획 소수자 텍스트마이닝: 토픽모델링 py2"
"""

import os
import pandas as pd

os.getcwd()
Path='C:\\Users\\hannews\\Documents\\2019 신년기획 소수자 텍스트마이닝'
os.chdir(Path)

# * load text *
# * reading one Excel file *
# 중복 제거는 raw data 단계에서 수행함
from openpyxl import load_workbook
from collections import defaultdict

Theme="gay"
RawFile='Dataset\\complete_gay.xlsx'
wb = load_workbook(RawFile)
ws = wb.active

# 키워드(칼럼 O)의 단어들 리스트로 읽기
words = ws['O']
words_list = [x.value.split(',') for x in words[1:]]

# 한번만 나타난 단어는 제거
frequency = defaultdict(int)
for text in words_list:
    for token in text:
        frequency[token] += 1
words_list = [[token for token in text if frequency[token] > 1] for text in words_list]

# 연도 데이터(칼럼 B) 읽기
year = ws['B']
year_list = [str(int(y.value))[:4] for y in year[1:]]
# end of one Excel file //

### * 토픽 모델링 * 
# * gensim LDA model *
import logging
logging.basicConfig(format = '%(asctime)s : %(levelname)s : %(message)s', level=logging.INFO)

from gensim import corpora
from gensim.models import LdaModel

# 여러 random_state 버전의 모델 만들기
random_range = list(range(4,11))
random_range = random_range[::-1]

K = 25 # the optimized number of the topics: gay 25
iterations = 1000 # 반복 횟수

while True:
    if not random_range:
        break
    random_state = random_range.pop() # set random seed
    fname="nor_lda_"+Theme+"K"+str(K)+"R"+str(random_state)+"I"+str(iterations)

    # build 토픽 모델
    dictionary = corpora.Dictionary(words_list)
    dictionary.save('TM\\'+fname+'_dictionary.pkl')

    corpus = [dictionary.doc2bow(text) for text in words_list]
    corpora.MmCorpus.serialize('TM\\'+fname+'_corpus.mm', corpus)

    lda = LdaModel(corpus, 
               num_topics=K, 
               id2word=dictionary,
               random_state=random_state, 
               iterations=iterations)
    lda.save('Models\\'+fname)


# 문서별 토픽의 비중을 토픽 당 비중으로 전환하는 함수
# convert document-weight to topic-weight
def convert_d2t(documents_topics):
    topic_weight = defaultdict(float)
    for document_topics in documents_topics:
        for topic in document_topics:
            topic_weight[topic[0]] += topic[1] / len(documents_topics) #normalize
    return topic_weight

# split corpus for each year
words_year = defaultdict(list)
for k, v in zip(year_list, words_list):
    words_year[k].append(v)    
    
## 여러 모델 결과 저장
random_range = list(range(4,11))
while True:
    if not random_range:
        break
    random_state = random_range.pop() # set random seed
    fname="nor_lda_"+Theme+"K"+str(K)+"R"+str(random_state)+"I"+str(iterations)

    dictionary = corpora.Dictionary.load('TM\\'+fname+'_dictionary.pkl')
    corpus = corpora.MmCorpus('TM\\'+fname+'_corpus.mm')
    lda = LdaModel.load('Models\\'+fname)
    topics = lda.print_topics(-1,20)

    # 토픽별 단어 저장
    feat_fname = 'TopicTrend\\'+fname+'_feats.txt'
    with open(feat_fname, 'w') as text_file:
        for topic_num, features in topics:
            text_file.write("Topic={0} \n {1} \n".format(topic_num, features))
    
    ### * Topic Trend *
    # * 연도별로 어떤 토픽이 어떤 비중을 차지하는가 *
    # bow of each year
    corpus_year = defaultdict(list)
    for k, texts in words_year.items():
        corpus_year[k] = [dictionary.doc2bow(text) for text in texts]
    
    # LDA vector of each year
    lda_year = defaultdict(list)
    for k, bows in corpus_year.items():
        lda_year[k] = [lda[bow] for bow in bows]

    # 연도를 key, 연도별 토픽 비중을 value로 하는 딕트로 변환
    keylist = list(lda_year.keys())
    keylist.sort()
    
    topic_year = defaultdict(dict)
    for key in keylist:
        topic_year[key]=convert_d2t(lda_year[key])
     
    # 파일로 저장
    TopicMatrixFilename="TopicTrend\\"+fname+"TM.csv"
    pd.DataFrame.from_dict(topic_year).to_csv(TopicMatrixFilename, index=False)
    # end of Topic Trend test //

## * visualization *
# !!주의!!
# pyLDAvis 패키지는 현재(19-08-19) python3.4 아래에서 돌아감 (p)
import pyLDAvis.gensim
import os
from gensim import corpora
from gensim.models import LdaModel

Path=u'C:\\Users\\hannews\\Documents\\2019 신년기획 소수자 텍스트마이닝'
os.chdir(Path)

Theme = 'gay'
K = 25 # the optimized number of the topics: gay 25
iterations = 1000 # 반복 횟수
random_range = list(range(4,11))

while True:
    if not random_range:
        break
    random_state = random_range.pop() # set random seed
    fname="nor_lda_"+Theme+"K"+str(K)+"R"+str(random_state)+"I"+str(iterations)

    dictionary = corpora.Dictionary.load('TM\\'+fname+'_dictionary.pkl')
    corpus = corpora.MmCorpus('TM\\'+fname+'_corpus.mm')
    lda = LdaModel.load('Models\\'+fname)
    prepared_data = pyLDAvis.gensim.prepare(lda, corpus, dictionary, sort_topics=False)
    # save as html file
    vname="LdaVis\\"+fname+".html" #sort_topics is False
    pyLDAvis.save_html(prepared_data, vname)